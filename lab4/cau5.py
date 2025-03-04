import openpyxl
import os
from tkinter import *
from tkinter import messagebox

# Định nghĩa đường dẫn file Excel
file_path = r"D:\Book1.xlsx"

# Nếu file chưa tồn tại, tạo file mới
if not os.path.exists(file_path):
    wb = openpyxl.Workbook()
    wb.save(file_path)

# Mở file
wb = openpyxl.load_workbook(file_path)
sheet = wb.active


# Hàm thiết lập Excel
def excel():
    columns = ["A", "B", "C", "D", "E", "F", "G"]
    headers = [
        "Name",
        "Course",
        "Semester",
        "Form Number",
        "Contact Number",
        "Email id",
        "Address",
    ]

    for col, header in zip(columns, headers):
        sheet.column_dimensions[col].width = 20
        sheet.cell(row=1, column=columns.index(col) + 1, value=header)

    wb.save(file_path)


# Hàm xóa nội dung nhập vào
def clear():
    for field in [
        name_field,
        course_field,
        sem_field,
        form_no_field,
        contact_no_field,
        email_id_field,
        address_field,
    ]:
        field.delete(0, END)


# Hàm lưu dữ liệu vào Excel
def insert():
    if all(
        field.get() == ""
        for field in [
            name_field,
            course_field,
            sem_field,
            form_no_field,
            contact_no_field,
            email_id_field,
            address_field,
        ]
    ):
        messagebox.showwarning("Input Error", "Please fill all fields!")
        return

    # Lưu dữ liệu vào Excel
    sheet.append(
        [
            name_field.get(),
            course_field.get(),
            sem_field.get(),
            form_no_field.get(),
            contact_no_field.get(),
            email_id_field.get(),
            address_field.get(),
        ]
    )

    wb.save(file_path)
    messagebox.showinfo("Success", "Data has been saved successfully!")
    clear()


# Tạo GUI
root = Tk()
root.configure(background="#f0f0f0")  # Màu nền sáng nhẹ
root.title("Registration Form")
root.geometry("500x450")  # Cải thiện kích thước của cửa sổ

# Chỉnh sửa font
label_font = ("Arial", 9)
entry_font = ("Arial", 10)

# Màu sắc hiện đại
button_color = "#5c6bc0"  # Màu tím nhạt cho nút
highlight_color = "#3f51b5"  # Màu xanh dương mạnh mẽ

# Tiêu đề
Label(
    root,
    text="Registration Form",
    bg="#f0f0f0",
    font=("Arial", 16, "bold"),
    fg=highlight_color,
).grid(row=0, column=1, pady=15)

# Các nhãn và ô nhập liệu
labels = [
    "Name",
    "Course",
    "Semester",
    "Form No.",
    "Contact No.",
    "Email id",
    "Address",
]
fields = []

for i, label in enumerate(labels):
    Label(root, text=label, bg="#f0f0f0", font=label_font, fg="#333").grid(
        row=i + 1, column=0, pady=8, padx=15
    )
    field = Entry(
        root,
        font=entry_font,
        relief="solid",
        bd=1,
        highlightthickness=1,
        highlightcolor=highlight_color,
    )
    field.grid(row=i + 1, column=1, ipadx="90", pady=5)
    fields.append(field)

(
    name_field,
    course_field,
    sem_field,
    form_no_field,
    contact_no_field,
    email_id_field,
    address_field,
) = fields

# Nút Submit
Button(
    root,
    text="Submit",
    fg="white",
    bg=highlight_color,
    font=("Arial", 10),
    relief="flat",
    command=insert,
).grid(row=8, column=1, pady=20, ipadx=8)

excel()  # Gọi hàm thiết lập Excel

root.mainloop()
